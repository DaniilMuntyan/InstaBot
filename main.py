import time
from datetime import time as time_, date as dt, datetime
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome import service
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment


def get_day(date):
    return date.strftime('%A')


def edit_date(date):
    s = date.split('T')
    yyyymmdd = s[0].split('-')
    ddmmyyyy = yyyymmdd[2] + "." + yyyymmdd[1] + "." + yyyymmdd[0]
    temp = s[1]
    my_date = ddmmyyyy.split('.')
    my_time = temp[:8].split(':')
    new_date = dt(int(my_date[2]), int(my_date[1]), int(my_date[0]))
    new_time = time_(int(my_time[0]), int(my_time[1]), int(my_time[2]))
    return new_date, new_time


def signIn(driver, email, password):
    driver.get("https://www.instagram.com/accounts/login/")
    time.sleep(1)
    emailInput = driver.find_elements_by_css_selector('form input')[0]
    passwordInput = driver.find_elements_by_css_selector('form input')[1]
    emailInput.send_keys(email)
    passwordInput.send_keys(password)
    passwordInput.send_keys(Keys.ENTER)
    return driver


def scroll_down(driver):
    last_height = driver.execute_script("return document.body.scrollHeight")
    my_list = []
    wait = WebDriverWait(driver, 1)
    while True:
        p = 0
        for i in range(1, 17):
            if p:
                break
            for j in range(1, 4):
                try:
                    s = '//div[@class=" _2z6nI"]/article[@class="ySN3v"]/div/div/div[%d]/div[%d]/a' % (i, j)
                    link = wait.until(EC.presence_of_element_located((By.XPATH, s)))
                    a = link.get_attribute('href')
                    if a not in my_list:
                        my_list.append(a)
                except (NoSuchElementException, TimeoutException):
                    p = 1
                    # print(str(i) + ' ' + str(j))
                    break

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height
    driver.execute_script("window.scrollTo(0, 0);")
    return my_list


# begin = "https://www.instagram.com/moonchan_daniil/"
# begin = "https://www.instagram.com/tatyanamuntyan/?hl=ru"
# begin = "https://www.instagram.com/tayinitskaya/"
'''k = ['https://www.instagram.com/tatyanamuntyan/?hl=ru', 
'https://www.instagram.com/mistake.away/?igshid=16iazcjlk44ma',
     'https://www.instagram.com/svetlana_tyshchenko/', 'https://www.instagram.com/bmiraw/']'''
'''k = ['https://www.instagram.com/serejikfeshchenko/',
     'https://www.instagram.com/krs.kl/', 'https://www.instagram.com/oleksii_stetsyk/',
     'https://www.instagram.com/_tommy__vercetti/']'''

k = ['https://www.instagram.com/oleksii_stetsyk/', 'https://www.instagram.com/moonchan_daniil/']

start_time = time.time()
# webdriver_service = service.Service('D:\chromedriver_win32\chromedriver')
'''webdriver_service = service.Service('D:\Python\Парсинг\Insta_bot\operadriver')
webdriver_service.start()
driver = webdriver.Remote(webdriver_service.service_url, webdriver.DesiredCapabilities.OPERA)'''

# Turn VPN on
opera_profile = r'C:\Users\Админ\AppData\Roaming\Opera Software\Opera Stable'
options = webdriver.ChromeOptions()
options.add_argument('user-data-dir=' + opera_profile)
options._binary_location = r'C:\Program Files (x86)\Opera\66.0.3515.72\opera.exe'
driver = webdriver.Opera(executable_path=r'D:\Python\Парсинг\Insta_bot\operadriver.exe', options=options)

# driver = signIn(driver, 'daniilmuntjan@gmail.com', '123456781')
for i in range(len(k)):
    begin = k[i]
    driver.get(begin)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Locations'
    wait = WebDriverWait(driver, 7)
    wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'XjzKX')))
    my_list = []

    w = driver.find_element_by_class_name('g47SY ').text

    print(w)

    try:
        ws.cell(row=1, column=1, value="№")
        ws.cell(row=1, column=2, value='Location')
        ws.cell(row=1, column=3, value='Site')
        ws.cell(row=1, column=4, value='Date')
        ws.cell(row=1, column=5, value='Time')
        ws.cell(row=1, column=6, value='Day')
        ws.cell(row=1, column=7, value='Likes')
        ws.cell(row=1, column=8, value='Views')

        c = 2

        # driver.get(begin)
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'XjzKX')))
        while len(my_list) is not int(w):
            my_list = scroll_down(driver)  # список ссылок на посты

        print(len(my_list))
        data_style = NamedStyle(name='data_style', number_format='dd.mm.yyyy')
        for i in range(0, len(my_list)):
            driver.get(my_list[i])
            flag = False
            t = WebDriverWait(driver, 7)
            curr = t.until(EC.presence_of_element_located((By.CLASS_NAME, 'JF9hh')))
            try:
                location = driver.find_element_by_class_name('O4GlU').text
            except NoSuchElementException:
                location = "-"

            insta_date = driver.find_element_by_tag_name('time').get_attribute('datetime')
            temp = ''
            isVideo = 0
            try:
                likes = driver.find_element_by_class_name('Nm9Fw').text
                temp = likes.split(' ')
            except NoSuchElementException:
                views = driver.find_element_by_class_name('vcOH2').text
                temp = views.split(' ')
                isVideo = 1
                temp[0] = temp[1]

            ws.cell(row=c, column=1, value=c-1)
            ws.cell(row=c, column=2, value=location)
            ws.cell(row=c, column=3, value='=HYPERLINK("{}")'.format(driver.current_url))
            ws.cell(row=c, column=4, value=edit_date(insta_date)[0])
            ws.cell(row=c, column=4).value = ws.cell(row=c, column=4).value.strftime('%d.%m.%Y')
            ws.cell(row=c, column=5, value=edit_date(insta_date)[1])
            ws.cell(row=c, column=6, value=get_day(edit_date(insta_date)[0]))
            ws.cell(row=c, column=6).style = data_style
            if not isVideo:
                ws.cell(row=c, column=7, value=int(temp[0]))
                ws.cell(row=c, column=8, value='-')
                ws.cell(row=c, column=8).alignment = Alignment(horizontal='right')
            else:
                ws.cell(row=c, column=7, value='-')
                ws.cell(row=c, column=8, value=int(temp[0]))
                ws.cell(row=c, column=7).alignment = Alignment(horizontal='right')

            c += 1

        font = Font(name='Times New Roman', size=16, bold=True)
        for j in range(1, len(my_list)):
            ws.cell(row=1, column=j).font = font
            ws.cell(row=1, column=j).alignment = Alignment(horizontal='center', vertical='center')

        font = Font(name='Times New Roman', size=14)
        for i in range(2, len(my_list) + 2):
            for j in range(1, 9):
                if j == 3:
                    ws.cell(row=i, column=j).font = Font(name='Times New Roman', color='0000FF',
                                                         underline='single', size=14)
                else:
                    ws.cell(row=i, column=j).font = font

        ws.column_dimensions['B'].width = 57
        ws.column_dimensions['C'].width = 52
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['H'].width = ws.column_dimensions['G'].width

    finally:
        g = begin.split('/')
        wb.save(g[len(g) - 2] + str(i) + '.xlsx')

driver.close()
driver.quit()
print("--- %s seconds ---" % (time.time() - start_time))
